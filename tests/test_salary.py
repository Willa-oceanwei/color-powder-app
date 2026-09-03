from pathlib import Path

from utils.database import DatabaseConfig, connect_from_config, initialize_database_with_health
from utils.salary_calculator import (calculate_leave_deduction, calculate_monthly_extra_totals,
                                     calculate_salary, generate_salary_note)
from utils.salary_excel import _monthly_summary, _payroll_leave_note
from utils.salary_repository import (annual_leave_balance_before_month, delete_salary,
                                     delete_annual_leave_history_record,
                                     get_annual_leave_setting, get_employee_salary_note,
                                     get_employee_salary_notes, get_month_salaries,
                                     get_salary_monthly_extras, get_settled_month_salaries,
                                     list_annual_leave_history, list_employees, list_settled_salaries_in_range,
                                     save_annual_leave_history_record, save_annual_leave_setting, save_employee, save_salary)
from utils.salary_repository import save_employee_salary_note, save_salary_monthly_extras


def test_salary_snapshot_survives_employee_raise(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "salary.db")
    initialize_database_with_health(config)
    employee = {"employee_id":"E001", "name":"王美文", "join_date":"2026-01-01", "active":1,
                "base_salary":29500, "attendance_bonus":2000, "cooling_allowance":500,
                "allowance":3100, "position_allowance":0, "insurance":1196,
                "standard_hours":8, "annual_leave_base":11,
                "special_addition_enabled":1, "special_addition_amount":1800,
                "special_addition_note":"每月預設", "default_deduction_enabled":1,
                "default_deduction_amount":500, "default_deduction_note":"借支", "note":""}
    save_employee(config, employee)
    snapshot = {"year":2026,"month":8,"employee_id":"E001","employee_name_snapshot":"王美文",
                "base_salary_snapshot":29500,"attendance_bonus_snapshot":2000,"cooling_allowance_snapshot":500,
                "allowance_snapshot":3100,"position_allowance_snapshot":0,"insurance_snapshot":1196,
                "standard_hours_snapshot":8,"leave_days":0,"leave_hours":0,"annual_leave_days":0,
                "annual_leave_hours":0,"annual_leave_balance_before":11,"annual_leave_balance_after":11,
                "late_deduction":0,"manual_note":"","system_note":""}
    snapshot.update(calculate_salary(snapshot))
    save_salary(config, snapshot, settle=True)
    save_employee(config, {**employee, "base_salary":32000})
    updated = list_employees(config)[0]
    assert updated["base_salary"] == 32000
    assert updated["special_addition_amount"] == 1800
    assert updated["default_deduction_note"] == "借支"
    assert get_month_salaries(config, 2026, 8)[0]["base_salary_snapshot"] == 29500


def test_mistaken_salary_can_be_deleted_without_deleting_employee(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "delete-salary.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E001", "name":"測試員工", "join_date":"2026-01-01"})
    data = {"year":2026, "month":7, "employee_id":"E001", "employee_name_snapshot":"測試員工"}
    data.update(calculate_salary(data))
    salary_id = save_salary(config, data, [{"type":"addition", "item_name":"特別加給", "amount":100, "note":"測試"}])
    assert get_settled_month_salaries(config, 2026, 7) == []

    delete_salary(config, salary_id)

    assert get_month_salaries(config, 2026, 7) == []
    assert list_employees(config)[0]["employee_id"] == "E001"
    with connect_from_config(config) as conn:
        assert conn.execute("SELECT is_deleted, deleted_at FROM salary_monthly WHERE salary_id=?", (salary_id,)).fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM salary_deletion_audit WHERE salary_id=?", (salary_id,)).fetchone()[0] == 1


def test_database_reinitialization_preserves_saved_salary_draft(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "persistent-draft.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E001", "name":"測試員工", "join_date":"2026-01-01"})
    draft = {"year":2026, "month":9, "employee_id":"E001", "employee_name_snapshot":"測試員工",
             "system_note":"保留自動備註", "manual_note":"保留人工備註"}
    draft.update(calculate_salary(draft))
    save_salary(config, draft, annual_leave_records=[
        {"date":"2026-09-03", "days":1, "hours":0, "note":"保留特休日期"},
    ])

    initialize_database_with_health(config)

    restored = get_month_salaries(config, 2026, 9)[0]
    assert restored["status"] == "draft"
    assert restored["system_note"] == "保留自動備註"
    assert restored["manual_note"] == "保留人工備註"
    assert restored["annual_leave_records"][0]["date"] == "2026-09-03"


def test_month_report_returns_only_settled_snapshots(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "settled-report.db")
    initialize_database_with_health(config)
    for employee_id, name in (("E1", "甲"), ("E2", "乙")):
        save_employee(config, {"employee_id":employee_id, "name":name, "join_date":"2026-01-01"})
        data = {"year":2026, "month":7, "employee_id":employee_id, "employee_name_snapshot":name}
        data.update(calculate_salary(data))
        save_salary(config, data, settle=employee_id == "E1")

    settled = get_settled_month_salaries(config, 2026, 7)
    assert [row["employee_id"] for row in settled] == ["E1"]


def test_personal_annual_leave_opening_balance_and_monthly_usage(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "annual-leave.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E1", "name":"甲", "join_date":"2026-01-01", "standard_hours":8})
    save_annual_leave_setting(config, "E1", 2026, 14, 9, 8, "歷年制特休14日")
    august = {"year":2026, "month":8, "employee_id":"E1", "employee_name_snapshot":"甲",
              "standard_hours_snapshot":8, "annual_leave_days":1, "annual_leave_hours":4,
              "annual_leave_entitlement_snapshot":14, "annual_leave_note_snapshot":"歷年制特休14日",
              "annual_leave_balance_before":9, "annual_leave_balance_after":7.5}
    august.update(calculate_salary(august))
    save_salary(config, august, settle=True)

    assert annual_leave_balance_before_month(config, "E1", 2026, 9) == 7.5
    saved = get_settled_month_salaries(config, 2026, 8)[0]
    assert saved["annual_leave_entitlement_snapshot"] == 14
    assert saved["annual_leave_balance_after"] == 7.5


def test_annual_leave_balance_falls_back_to_employee_current_days(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "annual-leave-fallback.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E1", "name":"甲", "join_date":"2026-01-01",
                           "standard_hours":8, "annual_leave_base":10})
    january = {"year":2026, "month":1, "employee_id":"E1", "employee_name_snapshot":"甲",
               "standard_hours_snapshot":8, "annual_leave_days":1, "annual_leave_hours":4,
               "annual_leave_balance_before":10, "annual_leave_balance_after":8.5}
    january.update(calculate_salary(january))
    save_salary(config, january, settle=True)

    assert get_annual_leave_setting(config, "E1", 2026) is None
    assert annual_leave_balance_before_month(config, "E1", 2026, 1) == 10
    assert annual_leave_balance_before_month(config, "E1", 2026, 2) == 10

    save_annual_leave_setting(config, "E1", 2026, 10, 10, 1, "")
    assert annual_leave_balance_before_month(config, "E1", 2026, 2) == 8.5


def test_leave_rounding_and_same_sheet_excel():
    import pytest
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from utils.salary_excel import generate_salary_workbook
    assert calculate_leave_deduction(29500, 1, 2, 30, 8) == 1229
    salary = {"employee_id":"E1","employee_name_snapshot":"甲","base_salary_snapshot":29500,"attendance_bonus_snapshot":3000,
              "cooling_allowance_snapshot":500,"allowance_snapshot":3000,"position_allowance_snapshot":1000,
              "leave_days":1,"leave_hours":2,"leave_deduction":1229,"insurance_snapshot":2168,
              "late_deduction":30,"final_salary":35373,
              "adjustments":[{"type":"addition","item_name":"特別加給","amount":1800,"note":"餐費"}],
              "system_note":"另有特別加給1,800元（餐費）。","manual_note":"8/15補發制服費", "standard_hours_snapshot":8,
              "annual_leave_days":1,"annual_leave_hours":4,"annual_leave_entitlement_snapshot":14,
              "annual_leave_balance_after":7.5,"annual_leave_note_snapshot":"歷年制特休14日"}
    from io import BytesIO
    first = {**salary, "company_cost_note":"甲公司負擔", "annual_leave_personal_note":"甲歷年制說明"}
    salaries = [first] + [
        {**salary, "employee_id":f"E{index}", "employee_name_snapshot":name,
         "company_cost_note":f"{name}公司負擔", "annual_leave_personal_note":f"{name}歷年制說明"}
        for index, name in enumerate(("乙", "丙", "丁", "戊"), 2)
    ]
    extras = {"employee_values":{"E1":30,"E2":0,"E3":10,"E4":0,"E5":5}, "previous_value":13873,
              "monthly_addition":30, "monthly_total":13903}
    workbook = load_workbook(BytesIO(generate_salary_workbook(2026, 8, salaries, extras)))
    assert workbook.sheetnames == ["2026年08月薪資"]
    sheet = workbook.active
    assert sheet.max_column == 13
    assert sheet["A1"].value == "2026年08月薪資"
    assert [cell.value for cell in sheet[2]] == ["月份／姓名", "底薪", "請假", "全勤", "涼水", "津貼",
                                                   "職務津貼", "勞健保", "遲到", "小計", "特別加給", "扣除額", "薪資總計"]
    assert [cell.value for cell in sheet[3]] == ["08月份 甲", 29500, -1229, 3000, 500, 3000, 1000,
                                                   -2168, -30, 33573, 1800, 0, 35373]
    assert sheet["A4"].value == "［特休1.5日，餘7.5日］\n［請假1.25日］\n8/15補發制服費"
    assert sheet["D4"].value == "公司負擔：\n甲公司負擔"
    assert "甲歷年制說明" in sheet["H4"].value
    assert sheet["A5"].value == "當月說明：上月13,873 + 本月新增30 = 本月總計13,903"
    assert "餐費" not in sheet["A4"].value
    assert sum(cell.value == "底薪" for row in sheet for cell in row) == 5
    assert any("乙" in str(cell.value) for row in sheet for cell in row)
    assert not any(cell.value == "歷年制特休" for row in sheet for cell in row)
    assert not any(cell.value == "每月附加數值區" for row in sheet for cell in row)
    assert sum(cell.value == "當月說明：上月13,873 + 本月新增30 = 本月總計13,903" for row in sheet for cell in row) == 5
    for coordinate in ("A4", "C4", "D4", "G4", "H4", "M4", "A5", "M5"):
        cell = sheet[coordinate]
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style in ("thin", "medium")
    assert all(sheet[cell].border.left.style == "thin" for cell in ("A4", "D4", "H4", "A5"))
    assert all(sheet[cell].border.right.style == "thin" for cell in ("C4", "G4", "M4", "M5"))
    assert sheet.max_row == 25
    assert sheet.sheet_properties.pageSetUpPr.fitToPage
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1


def test_excel_note_contains_leave_only():
    salary = {"leave_days":1, "leave_hours":2, "annual_leave_days":1, "annual_leave_hours":4,
              "standard_hours_snapshot":8, "annual_leave_balance_after":7.5,
              "system_note":"另有特別加給1,800元（餐費）。"}
    note = _payroll_leave_note(salary)
    assert note == "［特休1.5日，餘7.5日］\n［請假1.25日］"
    assert "餐費" not in note
    assert _payroll_leave_note({"annual_leave_days":1.2, "annual_leave_balance_after":4.8,
                                "standard_hours_snapshot":8}) == "［特休1.2日，餘4.8日］"
    assert _monthly_summary({"previous_value":13873, "monthly_addition":30, "monthly_total":13903}) == (
        "當月說明：上月13,873 + 本月新增30 = 本月總計13,903"
    )


def test_annual_leave_settings_are_scoped_by_employee_and_year(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "personal-annual-leave.db")
    initialize_database_with_health(config)
    for employee_id, name in (("E1", "甲"), ("E2", "乙")):
        save_employee(config, {"employee_id":employee_id, "name":name, "join_date":"2026-01-01"})
    save_annual_leave_setting(config, "E1", 2026, 14, 9, 8, "甲的2026年說明")
    save_annual_leave_setting(config, "E1", 2027, 15, 15, 1, "甲的2027年說明")
    save_annual_leave_setting(config, "E2", 2026, 10, 6, 8, "乙的2026年說明")

    assert get_annual_leave_setting(config, "E1", 2026)["note"] == "甲的2026年說明"
    assert get_annual_leave_setting(config, "E1", 2027)["annual_entitlement"] == 15
    assert get_annual_leave_setting(config, "E2", 2026)["note"] == "乙的2026年說明"


def test_personal_salary_notes_and_monthly_extras_are_editable(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "salary-notes.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E1", "name":"甲", "join_date":"2026-01-01"})
    save_employee_salary_note(config, "E1", 2026, "公司負擔A", "特休說明A")
    save_employee_salary_note(config, "E1", 2027, "公司負擔B", "特休說明B")
    save_employee_salary_note(config, "E1", 2026, "公司負擔A-更新", "特休說明A-更新")
    assert get_employee_salary_note(config, "E1", 2026)["company_cost_note"] == "公司負擔A-更新"
    assert get_employee_salary_note(config, "E1", 2027)["annual_leave_note"] == "特休說明B"
    assert get_employee_salary_notes(config, ["E1", "missing", "E1"], 2026)["E1"]["annual_leave_note"] == "特休說明A-更新"
    assert get_employee_salary_notes(config, [], 2026) == {}

    save_salary_monthly_extras(config, 2026, 7, {"E1":30}, 13873, 30, 13903)
    extras = get_salary_monthly_extras(config, 2026, 7)
    assert extras["employee_values"] == {"E1":30}
    assert extras["monthly_total"] == 13903


def test_monthly_extras_sum_employees_and_carry_previous_total():
    monthly_addition, monthly_total = calculate_monthly_extra_totals(
        13903, {"E1": 30, "E2": 12.5, "E3": 0},
    )
    assert monthly_addition == 42.5
    assert monthly_total == 13945.5


def test_generated_salary_note_includes_unique_annual_leave_dates():
    note = generate_salary_note({
        "annual_leave_days": 1,
        "annual_leave_hours": 2.5,
        "annual_leave_balance_after": 6.6875,
        "standard_hours_snapshot": 8,
        "annual_leave_records": [
            {"date": "2026-08-07"},
            {"date": "2026-08-07"},
            {"date": "2026-08-24"},
            {"date": ""},
        ],
    })
    assert "日期08/07、08/24" in note
    assert note.count("08/07") == 1


def test_dated_leave_records_are_linked_to_salary_and_editable(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "dated-leave.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E1", "name":"甲", "join_date":"2026-01-01", "standard_hours":8})
    data = {"year":2026,"month":8,"employee_id":"E1","employee_name_snapshot":"甲",
            "standard_hours_snapshot":8,"annual_leave_days":1.2,"annual_leave_hours":4,
            "annual_leave_balance_before":9,"annual_leave_balance_after":7.3,
            "system_note":"可修改的自動備註", "manual_note":"人工備註"}
    data.update(calculate_salary(data))
    save_salary(config, data, settle=True, annual_leave_records=[
        {"date":"2026-08-03","days":1,"hours":0,"note":""},
        {"date":"2026-08-12","days":0,"hours":4,"note":"下午特休"},
        {"date":"","days":0.2,"hours":0,"note":"未記日期"},
    ])
    records = list_annual_leave_history(config, "E1", 2026)
    assert [item["equivalent_days"] for item in records] == [1, 0.5, 0.2]
    assert records[-1]["date"] == ""
    assert all(item["salary_status"] == "settled" for item in records)

    # Saving the same salary again must not hide its existing dated records.
    save_salary(config, data, settle=True, annual_leave_records=records)
    assert len(list_annual_leave_history(config, "E1", 2026)) == 3
    resaved = get_settled_month_salaries(config, 2026, 8)[0]
    assert resaved["system_note"] == "可修改的自動備註"
    assert resaved["manual_note"] == "人工備註"

    save_annual_leave_history_record(config, {**records[1], "hours":2, "standard_hours":8})
    updated = list_annual_leave_history(config, "E1", 2026)
    assert updated[1]["equivalent_days"] == 0.25
    assert get_settled_month_salaries(config, 2026, 8)[0]["annual_leave_hours"] == 4
    delete_annual_leave_history_record(config, updated[0]["id"])
    assert len(list_annual_leave_history(config, "E1", 2026)) == 2


def test_annual_leave_batch_editor_normalizes_rows_and_ignores_empty_rows():
    import pytest
    pd = pytest.importorskip("pandas")
    from utils.salary_ui import (_annual_leave_editor_key, _annual_leave_editor_rows,
                                 _annual_leave_records_from_editor)

    existing = [{"date":"2026-08-03", "days":1, "hours":0, "note":"上午"},
                {"date":"", "days":0, "hours":2, "note":""}]
    editor_rows = _annual_leave_editor_rows(existing)
    assert pd.isna(editor_rows.loc[0, "時數"])
    assert editor_rows.loc[1, "時數"] == 2.0
    assert _annual_leave_editor_key("2026-08", 0, "E1") != _annual_leave_editor_key("2026-08", 0, "E2")
    editor_rows.loc[len(editor_rows)] = [None, None, None, None]

    assert _annual_leave_records_from_editor(editor_rows, 2026) == existing

    editor_rows.loc[1, "時數"] = 1.25
    assert _annual_leave_records_from_editor(editor_rows, 2026)[1]["hours"] == 1.25

    editor_rows.loc[0, "日期（可留空）"] = "8/3"
    assert _annual_leave_records_from_editor(editor_rows, 2026)[0]["date"] == "2026-08-03"
    editor_rows.loc[0, "日期（可留空）"] = "0824"
    assert _annual_leave_records_from_editor(editor_rows, 2026)[0]["date"] == "2026-08-24"
    editor_rows.loc[0, "日期（可留空）"] = "20260807"
    assert _annual_leave_records_from_editor(editor_rows, 2026)[0]["date"] == "2026-08-07"
    editor_rows.loc[0, "日期（可留空）"] = "2/30"
    with pytest.raises(ValueError, match="不是有效日期"):
        _annual_leave_records_from_editor(editor_rows, 2026)


def test_generated_salary_note_refreshes_until_user_edits_it():
    import pytest
    pytest.importorskip("pandas")
    from utils.salary_ui import _sync_generated_note_state

    state = {}
    _sync_generated_note_state(state, "note_E1", "自動內容一")
    assert state["note_E1"] == "自動內容一"

    _sync_generated_note_state(state, "note_E1", "自動內容二")
    assert state["note_E1"] == "自動內容二"

    state["note_E1"] = "自行修改內容"
    _sync_generated_note_state(state, "note_E1", "自動內容三")
    assert state["note_E1"] == "自行修改內容"

    saved_state = {}
    _sync_generated_note_state(saved_state, "note_E2", "重新計算內容", "已儲存的編輯內容")
    assert saved_state["note_E2"] == "已儲存的編輯內容"


def test_duplicate_salary_blocks_keep_only_first_employee_entry():
    import pytest
    pytest.importorskip("pandas")
    from utils.salary_ui import _deduplicate_salary_blocks

    first = {"employee_id": "E1", "manual_note": "保留這筆"}
    duplicate = {"employee_id": "E1", "manual_note": "刪除這筆"}
    second = {"employee_id": "E2"}
    assert _deduplicate_salary_blocks([first, duplicate, second]) == [first, second]


def test_draft_report_rows_include_employee_notes(tmp_path: Path):
    import pytest
    pytest.importorskip("pandas")
    from utils.salary_ui import _salary_report_rows

    config = DatabaseConfig("sqlite", tmp_path / "draft-preview.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E1", "name":"甲", "join_date":"2026-01-01"})
    save_employee_salary_note(config, "E1", 2026, "公司負擔", "個人特休說明")

    rows, missing = _salary_report_rows(
        config, [{"employee_id":"E1", "employee_name_snapshot":"甲"}], 2026,
    )

    assert rows[0]["company_cost_note"] == "公司負擔"
    assert rows[0]["annual_leave_personal_note"] == "個人特休說明"
    assert missing == []


def test_salary_total_range_uses_only_settled_active_snapshots(tmp_path: Path):
    config = DatabaseConfig("sqlite", tmp_path / "salary-range.db")
    initialize_database_with_health(config)
    save_employee(config, {"employee_id":"E1", "name":"甲", "join_date":"2026-01-01"})
    salary_ids = {}
    for month, settled in ((1, True), (2, False), (3, True), (8, True)):
        data = {"year":2026,"month":month,"employee_id":"E1","employee_name_snapshot":"甲",
                "base_salary_snapshot":30000}
        data.update(calculate_salary(data))
        salary_ids[month] = save_salary(config, data, settle=settled)
    delete_salary(config, salary_ids[3])

    rows = list_settled_salaries_in_range(config, "E1", 2026, 1, 2026, 7)
    assert [(item["year"], item["month"]) for item in rows] == [(2026, 1)]
