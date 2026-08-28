"""Generate a compact, landscape, single-sheet monthly payroll report."""
from io import BytesIO


PAYROLL_HEADERS = (
    "月份／姓名", "底薪", "請假", "全勤", "涼水", "津貼", "職務津貼",
    "勞健保", "遲到", "小計", "特別加給", "扣除額", "薪資總計",
)


def _adjustment_total(salary, adjustment_type):
    return sum(
        int(item.get("amount") or 0)
        for item in salary.get("adjustments", [])
        if item.get("type") == adjustment_type
    )


def _leave_used_days(salary):
    hours_per_day = float(salary.get("standard_hours_snapshot") or 8)
    return float(salary.get("annual_leave_days") or 0) + float(salary.get("annual_leave_hours") or 0) / hours_per_day


def _payroll_leave_note(salary):
    """Build the compact monthly leave text shown in the slip's A:C note block."""
    parts = []
    annual_used = _leave_used_days(salary)
    if annual_used:
        balance = float(salary.get("annual_leave_balance_after") or 0)
        parts.append(f"［特休{annual_used:g}日，餘{balance:g}日］")
    hours_per_day = float(salary.get("standard_hours_snapshot") or 8)
    leave_used = float(salary.get("leave_days") or 0) + float(salary.get("leave_hours") or 0) / hours_per_day
    if leave_used:
        parts.append(f"［請假{leave_used:g}日］")
    return "\n".join(parts)


def _format_extra_number(value):
    number = float(value or 0)
    if number.is_integer():
        return f"{number:,.0f}"
    return f"{number:,.2f}".rstrip("0").rstrip(".")


def _monthly_summary(monthly_extras):
    return (
        f"當月說明：上月{_format_extra_number(monthly_extras.get('previous_value', 0))}"
        f" + 本月新增{_format_extra_number(monthly_extras.get('monthly_addition', 0))}"
        f" = 本月總計{_format_extra_number(monthly_extras.get('monthly_total', 0))}"
    )


def generate_salary_workbook(year, month, salaries, monthly_extras=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    salaries = list(salaries)
    monthly_extras = monthly_extras or {}
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month:02d}月薪資"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="777777")
    medium = Side(style="medium", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Every employee receives the same five-row slip: header, values, personal
    # notes, monthly leave note, and one blank separator.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PAYROLL_HEADERS))
    title = ws.cell(1, 1, f"{year}年{month:02d}月薪資")
    title.font = Font(name="Microsoft JhengHei", size=14, bold=True)
    title.alignment = centered
    ws.row_dimensions[1].height = 25
    row = 2
    for salary in salaries:
        for column, header in enumerate(PAYROLL_HEADERS, 1):
            cell = ws.cell(row, column, header)
            cell.font = Font(name="Microsoft JhengHei", size=9, bold=True)
            cell.alignment = centered
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[row].height = 28
        additions = _adjustment_total(salary, "addition")
        deductions = _adjustment_total(salary, "deduction")
        final_salary = int(salary.get("final_salary") or 0)
        subtotal = final_salary - additions + deductions
        values = (
            f"{month:02d}月份 {salary['employee_name_snapshot']}",
            int(salary.get("base_salary_snapshot") or 0),
            -int(salary.get("leave_deduction") or 0),
            int(salary.get("attendance_bonus_snapshot") or 0),
            int(salary.get("cooling_allowance_snapshot") or 0),
            int(salary.get("allowance_snapshot") or 0),
            int(salary.get("position_allowance_snapshot") or 0),
            -int(salary.get("insurance_snapshot") or 0),
            -int(salary.get("late_deduction") or 0),
            subtotal,
            additions,
            deductions,
            final_salary,
        )
        for column, value in enumerate(values, 1):
            cell = ws.cell(row + 1, column, value)
            cell.font = Font(name="Microsoft JhengHei", size=9, bold=column == 13)
            cell.alignment = centered
            cell.border = border
            if column > 1:
                cell.number_format = '#,##0;[Red]-#,##0;0'
            if column in (10, 13):
                cell.fill = total_fill
        ws.row_dimensions[row + 1].height = 24

        ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=3)
        ws.merge_cells(start_row=row + 2, start_column=4, end_row=row + 2, end_column=7)
        ws.merge_cells(start_row=row + 2, start_column=8, end_row=row + 2, end_column=13)
        # Apply each merged range's outer border after merging. This avoids the
        # fragmented borders that openpyxl can otherwise leave on merged cells.
        for start, end in ((1, 3), (4, 7), (8, 13)):
            for column in range(start, end + 1):
                ws.cell(row + 2, column).border = Border(
                    left=thin if column == start else None,
                    right=thin if column == end else None,
                    top=thin, bottom=thin,
                )
        automatic_note = _payroll_leave_note(salary)
        manual_note = str(salary.get("manual_note") or "").strip()
        ws.cell(row + 2, 1, "\n".join(part for part in (automatic_note, manual_note) if part))
        company_text = str(salary.get("company_cost_note") or "").strip()
        ws.cell(row + 2, 4, f"公司負擔：\n{company_text}" if company_text else "公司負擔：")
        annual_text = salary.get("annual_leave_personal_note") or salary.get("annual_leave_note_snapshot") or ""
        ws.cell(row + 2, 8, f"歷年制特休：\n{annual_text}" if annual_text else "歷年制特休：")
        for start in (1, 4, 8):
            cell = ws.cell(row + 2, start)
            cell.font = Font(name="Microsoft JhengHei", size=8, bold=False)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row + 2].height = 64

        ws.merge_cells(start_row=row + 3, start_column=1, end_row=row + 3, end_column=13)
        for column in range(1, 14):
            ws.cell(row + 3, column).border = Border(
                left=thin if column == 1 else None,
                right=thin if column == 13 else None,
                top=thin, bottom=medium,
            )
        note_cell = ws.cell(row + 3, 1, _monthly_summary(monthly_extras))
        note_cell.font = Font(name="Microsoft JhengHei", size=8)
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row + 3].height = 22
        ws.row_dimensions[row + 4].height = 9
        row += 5

    widths = (18, 11, 10, 10, 10, 11, 12, 11, 9, 12, 12, 11, 13)
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width

    ws.print_area = f"A1:M{max(row, 1)}"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15)
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
